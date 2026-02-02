#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Eleanor Megeath

CS 111

Spring 2025

Final Project
"""

def equilibriumOutput():
    """
    Calculates the output of an economy in equilibrium based on a serires of inputs
    
    Parameters
    ----------
    None
    
    Returns
    -------
    Y: equilibrium output, int
    
    """
    C0 = float(input("Autonomous Consumption: "))
    C1 = float(input("Propensity to Consume: "))
    I = float(input("Investment: "))
    G = float(input("Government Spending: "))
    X = float(input("Exports: "))
    IM = float(input("Imports: "))
    TType = input("Type of Taxes: ")
    if TType == "lump sum" or TType == "Lump Sum":
        T = float(input("Total Taxes: "))
        Y = (1/(1 - C1)) * (C0 - C1 * T + I + G + X - IM)
    elif TType == "proportional to income" or TType == "Proportional to Income":
        T = float(input("Tax Rate: "))
        Y = (1/(1 + C1 * T - C1)) * (C0 + I + G + X - IM)
    return int(Y)

        
def equilibriumEconomyValues(Y, C0, C1, I, G, X, IM, TType, T):
    """
    Calculates a value for economic models based on other values
    
    Parameters 
    ----------
    Y: output, int
    C0: autonomous consumption, int
    C1: propensity to consume, int
    I: investment, int
    G: government spending, int
    X: exports, int
    IM: imports, int
    TType: if taxes are proportional to income or lump sum, str
    T: total taxes or tax rate, int
    
    Returns
    -------
    Y, C0, C1, I, G, X, IM, or T; int
    
    """
    if Y == "Y":
        if TType == "lump sum" or TType == "Lump Sum":
            Y = (1/(1 - C1)) * (C0 - C1 * T + I + G + X - IM)
            return int(Y)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            Y = (1/(1 + C1 * T - C1)) * (C0 + I + G + X - IM)
            return int(Y)
    elif C0 == "C0":
        if TType == "lump sum" or TType == "Lump Sum":
            C0 = C1 * T - I - G - X + IM - Y * (1 - C1)
            return int(C0)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            C0 = Y * (1 + C1 * T - C1) - I - G - X + IM
            return int(C0)
    elif C1 == "C1":
        if TType == "lump sum" or TType == "Lump Sum":
            C1 = (1 / (T - Y)) * (C0 - Y + I + G + X - IM)
            return int(C1)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            C1 = (1 / T - 1) * ((1 / Y) * (C0 + I + G + X - IM) - 1)
            return int(C1)
    elif I == "I":
        if TType == "lump sum" or TType == "Lump Sum":
            I = C1 * T - C0 - G - X + IM - Y * (1 - C1)
            return int(I)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            I = Y * (1 + C1 * T - C1) - C0 - G - X + IM
            return int(I)
    elif G == "G":
        if TType == "lump sum" or TType == "Lump Sum":
            G = C1 * T - C0 - I - X + IM - Y * (1 - C1)
            return int(G)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            G = Y * (1 + C1 * T - C1) - C0 - I - X + IM
            return int(G)
    elif X == "X":
        if TType == "lump sum" or TType == "Lump Sum":
            X = C1 * T - C0 - I - G + IM - Y * (1 - C1)
            return int(X)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            X = Y * (1 + C1 * T - C1) - C0 - I - G + IM
            return int(X)
    elif IM == "IM":
        if TType == "lump sum" or TType == "Lump Sum":
            IM = C0 - C1 * T + I + G + X + Y(1 - C1)
            return int(IM)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            IM = C0 - Y(1 + C1 * T - C1) + I + G + X
            return int(IM)
    elif T == "T":
        if TType == "lump sum" or TType == "Lump Sum":
            T = (1 / C1) * (C0 - Y * (1 - C1) + I + G + X - IM)
            return int(T)
        elif TType == "proportional to income" or TType == "Proportional to Income":
            T = (1 / C1) * ((1 / Y) * (C0 + I + G + X - IM) - 1 + C1)
            return int(T)


import matplotlib.pyplot as pyplot 

def rounding(num, decimalPlaces):
    """
    Rounds numbers, used in outputAndDemand function
    
    Parameters
    ----------
    num: number to be rounded, float
    decimalPlaces: number of decimal places to be rounded to, int 
    
    Returns
    -------
    num: rounded number, float
    
    """
    if type(num) == float:
        num = str(num)
        for i in range(len(num)):
            if num[i] == ".":
                num = num[:i + decimalPlaces + 1]
                num = float(num)
                return num
    else:
        return num
    
    
def outputAndDemand(Z, C0, C1, T, I, G, timeSpan, timeChange, change):
    """
    Graphs and prints a table of output and demand in a hypothetical economy, the model assumes that net exports is equal to zero

    Parameters
    ----------
    Z: demand, int
    C0: autonomous consumption, int
    C1: propensity to consume, int
    T: lump sum taxes, int
    I: investment, int
    G: government spending, int
    timeSpan: number of periods graphed, int
    timeChange: period in which change occurs, int
    change: variable (Z, C0, C1, T, I, or G) that changes and its new value separated by a colon, str

    Returns
    -------
    Table of the rounded demand and output of each period
    
    """
    # initialize variables for graph
    timeList = []
    ZList = []
    YList = []
    
    time = 0
    
    # initialize variables to hold the variable and value changed
    cleanChangeVariable = ""                
    cleanChangeValue = ""
    
    # separate change into its variable and value
    for index in range(len(change)):   
        if ord(change[index]) == 58:
            changeVariable = change[:index]
            changeValue = change[index + 1:] 
        
    # remove any characters from changeVariable and changeValue that are not letters or numbers
    for index in range(len(changeVariable)):
        if (ord(changeVariable[index]) >= 48 and ord(changeVariable[index]) <= 57) or (ord(changeVariable[index]) >= 65 and ord(changeVariable[index]) <= 90):     
            cleanChangeVariable = cleanChangeVariable + changeVariable[index]
            
    for index in range(len(changeValue)):
        if ord(changeValue[index]) >= 48 and ord(changeValue[index]) <= 57:
            cleanChangeValue = cleanChangeValue + changeValue[index]     
    
    cleanChangeValue = int(cleanChangeValue)
    
    # initialize the starting value of demand
    Zc = Z

    for i in range(timeChange):
        Yc = Zc
        Zc = C0 + C1 * (Yc - T) + I + G
        time = time + 1
        
        YList.append(Yc)
        ZList.append(Zc)
        timeList.append(time)
                
    for i in range(timeChange, timeSpan):
        if cleanChangeVariable == "C0":
            Yc = Zc
            Zc = cleanChangeValue + C1 * (Yc - T) + I + G
            time = time + 1
    
            YList.append(Yc)
            ZList.append(Zc)
            timeList.append(time)
            
        elif cleanChangeVariable == "C1":
            Yc = Zc
            Zc = C0 + cleanChangeValue * (Yc - T) + I + G
            time = time + 1
            
            YList.append(Yc)
            ZList.append(Zc)
            timeList.append(time)
            
        elif cleanChangeVariable == "T":
            Yc = Zc
            Zc = C0 + C1 * (Yc - cleanChangeValue) + I + G
            time = time + 1
            
            YList.append(Yc)
            ZList.append(Zc)
            timeList.append(time)
            
        elif cleanChangeVariable == "I":
            Yc = Zc
            Zc = C0 + C1 * (Yc - T) + cleanChangeValue + G
            time = time + 1
            
            YList.append(Yc)
            ZList.append(Zc)
            timeList.append(time)
            
        elif cleanChangeVariable == "G":
            Yc = Zc
            Zc = C0 + C1 * (Yc - T) + I + cleanChangeValue
            time = time + 1
            
            YList.append(Yc)
            ZList.append(Zc)
            timeList.append(time)
    
    # for scale of graph
    zeroList = []
    for i in range(timeSpan):
        zeroList.append(0)       
    
    # graphing
    pyplot.plot(timeList, YList, label = "Output")
    pyplot.plot(timeList, ZList, label = "Demand")
    pyplot.plot(timeList, zeroList, alpha = 0)
    pyplot.legend(loc = 'center right')
    pyplot.xlabel('Time Period')
    pyplot.show()
    
    # table
    for i in range(len(YList)):
        YList[i] = rounding(YList[i], 3)
    for i in range(len(ZList)):
        ZList[i] = rounding(ZList[i], 3)
        
    matrix = [timeList, YList, ZList]
    for row in range(len(matrix)):
        for col in range(len(matrix[row])):
            if type(matrix[row][col]) == float:
                print(matrix[row][col], "\t", end = "")
            else:
                print(matrix[row][col], "\t", "\t", end = "")
        print()
             

def smoothingData(data, width):
    """
    Smooth out the trend of a list

    Parameters
    ----------
    data: list of integers
    width: window used to calculate the mean, int

    Returns
    -------
    smoothedData: list of averages, int 

    """
    smoothedData = []
    total = 0
    for i in range(width):
        total = total + data[i]
    
    for i in range(len(data)):
        width = min(width, len(data) - i)
        smoothedData.append(total / width)
        total = total - data[i]
        if i + width < len(data):
            total = total + data[i + width]
            
    return smoothedData


import openpyxl        

def graphingData():
    """
    Graphs real gdp per capita and population of France over time from Excel file
    
    Parameters
    ----------
    None.

    Returns
    -------
    None.

    """
    data = openpyxl.load_workbook("CSFinalData.xlsx")
    data_gdp = data["rgdpnapc"]
    data_pop = data["pop"]
    
    gdpList = [data_gdp.cell(row = i, column = 53).value for i in range(547, 648)]
    
    popList = [data_pop.cell(row = i, column = 53).value for i in range(547, 648)]
    
    time = 1819
    timeList = []
    for i in range(101):
        time = time + 1
        timeList.append(time)
        
    # inital graph
    pyplot.plot(popList, gdpList, "o")
    pyplot.xlabel("Population")
    pyplot.ylabel("Real GDP per Capita in 2011 Prices")
    pyplot.title("Real GDP Per Capita vs Population in France from 1820 to 1920")
    pyplot.show()
    
    pyplot.plot(timeList, gdpList)
    pyplot.xlabel("Year")
    pyplot.ylabel("Real GDP per Capita in 2011 Prices")
    pyplot.title("Real GDP per Capita in France over time")
    pyplot.show()
    
    # smooth gdp graph
    smoothGDP = smoothingData(gdpList, 5)
    
    pyplot.plot(timeList, smoothGDP)
    pyplot.xlabel("Population")
    pyplot.ylabel("Real GDP per Capita in 2011 Prices")
    pyplot.title("Real GDP Per Capita in France from 1820 to 1920")
    pyplot.show()


import random 

def expectedInflation(timeSpan, predictedInflation, m, z, a, theta):
    """
    Generate either a Phillips Curve graph or Accelerationist Phillips Curve based on expected inflation as well as model graphs showing the perfect Phillips Curve and Accelerationist Phillips Curve Relationships

    Parameters
    ----------
    timeSpan: number of time periods, int
    predictedInflation: value of predicted rate of inflation given by the Federal Bank, int
    m: markup of price over labor cost, int
    z: other factors affecting labor market equilibrium, int
    a: strenth of the relationship between unemployment and inflation, int
    theta: weight connecting expected and actual inflation with 0 being fully achnored inflation expectations and 1 being fully de-anchored expectations, str or int

    Returns
    -------
    None

    """
    uList = [random.randrange(3, 7) + rounding(random.random(), 1)]
    timeList = []
    inflationList = []
    changeInflationList = [0]
    
    for i in range(1, timeSpan):
        uChange = random.random()
        if uChange >= 0.5:
            uList.append(uList[i - 1] + rounding(random.randrange(0, 2) + random.random(), 1))
        else:
            uList.append(uList[i - 1] - rounding(random.randrange(0, 2) + random.random(), 1))
        timeList.append(i + 1)
    
    if theta == "random" or theta == "Random":
        theta = random.uniform(0, 1)
        print(theta)
    
    for i in range(timeSpan):
        if i < 1:
            inflation = predictedInflation + (m + z) - a * uList[i]
        else:
            inflation = (1 - theta) * predictedInflation + theta * inflationList[i - 1] + (m + z) - a * uList[i]
        inflationList.append(rounding(inflation, 3))
    
    if theta < 0.5:
        pyplot.plot(uList, inflationList, "o")
        pyplot.xlabel("Unemployment Rate")
        pyplot.ylabel("Inflation Rate")
        pyplot.title("Phillips Curve")
        pyplot.show()
        
        if theta < 0.2:
            print("Inflation expectations are anchored")
        
    else:
        for i in range(1, timeSpan):
            changeInflation = inflationList[i] - inflationList[i - 1] 
            changeInflationList.append(changeInflation)
        
        pyplot.plot(uList, changeInflationList, "o")
        pyplot.xlabel("Unemployment Rate")
        pyplot.ylabel("Change in the Inflation Rate")
        pyplot.title("Accelerationist Phillips Curve")
        pyplot.show()
        
        if theta > 0.8:
            print("Inflation expectations are de-anchored")
          
        # model graphs
    inflationAList = []
    inflationDList = []
    changeInflationDList = [0]
        
    for i in range(timeSpan):
        inflationA = predictedInflation + (m + z) - a * uList[i]
        inflationAList.append(rounding(inflationA, 3))
            
        if i < 1:
            inflationDList.append(inflationAList[0])
        else:
            inflationD = inflationDList[i - 1] + (m + z) - a * uList[i]
            inflationDList.append(rounding(inflationD, 3))
            changeInflationDList.append(inflationDList[i] - inflationDList[i - 1])
   
    pyplot.plot(uList, inflationAList, "o")
    pyplot.xlabel("Unemployment Rate")
    pyplot.ylabel("Inflation Rate")
    pyplot.title("Model Phillips Curve")
    pyplot.show()
    
    pyplot.plot(uList, changeInflationDList, "o")
    pyplot.xlabel("Unemployment Rate")
    pyplot.ylabel("Change in the Inflation Rate")
    pyplot.title("Model Accelerationist Phillips Curve")
    pyplot.show()
    
    
def main():
    # funtion one
    print(equilibriumOutput())
    # ex: 4, 0.8, 8, 12, 0, 0, lump sum, and 18 will give you 48
    
    # function two
    print(equilibriumEconomyValues(2727, 200, 0.5, 1000, "G", 0, 0, "proportional to income", 0.1))
    
    # function three
    outputAndDemand(64, 4, 0.8, 14, 8, 12, 10, 2, "T : 10")
    
    # function four
    graphingData()        

    # function five
    expectedInflation(30, 3, 0.5, 0.5, 0.2, "random")
    
main()

